from urllib.request import urlopen
from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import *
import xlrd
import sys
import os

driver = webdriver.Chrome()

def open_excel():
    excel = load_workbook("vaskemaskin_data.xlsx")
    excel_ark = excel["vaskemaskin_data"]
    excel1 = xlrd.open_workbook("vaskemaskin_data.xlsx")
    excel_ark1 = excel1.sheet_by_name("vaskemaskin_data")
    return excel,excel_ark,excel_ark1
def klokke_maskin(hey):
    klokke,dag = klokke_dag()
    x=int(klokke)
    if klokke == 0:
        x=24

    if dag == "mandag":
        y=2+hey
    elif dag == "tirsdag":
        y=22+hey
    elif dag == "onsdag":
        y=42+hey
    elif dag == "torsdag":
        y=63+hey
    elif dag == "fredag":
        y=83+hey
    elif dag == "lørdag":
        y=103+hey
    elif dag == "søndag":
        y=123+hey
    else:
        print("Error in day time function")
        sys.exit()
    
    if dag == "mandag":
        y1=0
    elif dag == "tirsdag":
        y1=20
    elif dag == "onsdag":
        y1=40
    elif dag == "torsdag":
        y1=60
    elif dag == "fredag":
        y1=81
    elif dag == "lørdag":
        y1=101
    elif dag == "søndag":
        y1=121
    else:
        print("Error in day time function_2")
        sys.exit()
    return x,y,y1
def klokke_dag():
    try:
        link = urlopen("https://www.timeanddate.no/klokka/i/norge/kristiansand")
        klokke=[]
        alt = BeautifulSoup(link, 'html' 'features="lxml"')
        klokke = alt.find_all('span')
        span_dag = klokke[4]
        span1_dag = str(span_dag)
        span2_dag = span1_dag.split(">")
        span3_dag = span2_dag[1].split(" ")
        dag = span3_dag[0] 
        span_klokke = klokke[2]
        span_klokke1 = str(span_klokke)
        klokke1 = span_klokke1.split("<")
        klokke2 = klokke1[1]
        f_klokke = klokke2.split(">")
        ff_klokke = f_klokke[1]
        fff_klokke = ff_klokke.split(":")
        klokke = fff_klokke[0]
        return klokke,dag
    except:
        driver.quit()
        sys.exit()
def add_tall(liste,maskin_nr,x,y,y1):
    excel,excel_ark,excel_ark1 = open_excel()
    count=0
    count = excel_ark1.cell_value(y1,x)
    if count == "":
        count=0
    int(count)
    count+=1
    excel_ark.cell(y1+1,x+1,count)
    if liste[0] == "Idle" or liste[0] == "Closing":
        verdi=0
        verdi = excel_ark1.cell_value(y,x)
        if verdi == "":
            verdi=0
        int(verdi)
        verdi+=1
        excel_ark.cell(y+1,x+1,verdi)
    excel.save("vaskemaskin_data.xlsx")
def open_webpage_and_run():
    maskin_liste = []
    link = "https://mielelogic.com/#/laundrystate"
    alle_maskiner =[
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[1]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[2]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[3]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[4]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[5]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[6]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[7]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[8]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[9]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[10]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[11]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[12]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[13]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[14]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[15]/div/div[5]',
        '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[16]/div/div[5]'  
        ]
    
    driver.get(link)
    sleep(5)            # SE HER HVIS DU FÅR ERROR!
    try:
        click_norge = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[1]/div/div/button')
        click_norge.click()
        click_norge = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[1]/div/div/ul/li[2]/a/span[1]')
        click_norge.click()
        brukernavn = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[2]/div/div/input')
        brukernavn.click()
        brukernavn.send_keys("My username")
        passord = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[3]/div[1]/div[1]/input')
        passord.click()
        passord.send_keys("My password")
        login = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[4]/button')
        login.click()
        sleep(5)          # OG HER
        vaskeri_status = driver.find_element_by_xpath('//*[@id="content"]/div/div/div/div/div[3]/div/div[1]/a/h4')
        sleep(5)          # OG HER
        vaskeri_status.click()
        sleep(10)         # OG HER
        hey=0
        for e in alle_maskiner:    
            for i in driver.find_elements_by_xpath(e):
                maskin_liste.append(i.text)
            maskin_1 = maskin_liste[hey].split(" ")
            print(maskin_1[0])
            print("1")
            x,y,y1 = klokke_maskin(hey)
            add_tall(maskin_1,hey,x,y,y1)
            hey+=1
    except:
        print("Error in open_webpage_and_run")
        driver.quit()
        sys.exit()
def main():
    open_webpage_and_run()
    driver.quit()
    sys.exit()
main()
