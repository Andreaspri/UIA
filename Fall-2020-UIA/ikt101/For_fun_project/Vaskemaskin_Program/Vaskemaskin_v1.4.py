from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
import datetime
import xlrd
import sys
import os

error_log_path = "/home/pi/Desktop/Vaskemaskin_Program"
excel_fil = "/home/pi/Desktop/Vaskemaskin_Program/vaskemaskin_data.xlsx"

driver = webdriver.Chrome()

def open_excel():
    excel = load_workbook(excel_fil)
    excel_ark = excel["vaskemaskin_data"]
    excel1 = xlrd.open_workbook(excel_fil)
    excel_ark1 = excel1.sheet_by_name("vaskemaskin_data")
    return excel,excel_ark,excel_ark1

def klokke_maskin(count,klokke,dag):
    x=int(klokke)
    if klokke == 0 or klokke == 00:
        x=24

    if dag == "Monday":
        y=2+count
        y1=0
    elif dag == "Tuesday":
        y=22+count
        y1=20
    elif dag == "Wednesday":
        y=42+count
        y1=40
    elif dag == "Thursday":
        y=63+count
        y1=60
    elif dag == "Friday":
        y=83+count
        y1=81
    elif dag == "Saturday":
        y=103+count
        y1=101
    elif dag == "Sunday":
        y=123+count
        y1=121
    else:
        klokke,dato = dato_tid()
        os.system(f"cd {error_log_path} && {dato} {klokke} echo Error assigning x,y corndinates >> error_log.txt")
        driver.quit()
        sys.exit()
    return x,y,y1

def dato_tid():
    klokke = str(datetime.datetime.now().time()).split('.')
    dato = str(datetime.datetime.now()).split(" ")
    return klokke[0],dato[0]

def klokke_dag():
    try:
        dag = str(datetime.datetime.today().strftime("%A"))
        klokke = str(datetime.datetime.now().time()).split(":")
        return int(klokke[0]),dag
    except:
        klokke,dato = dato_tid()
        os.system(f"cd {error_log_path} && echo {dato} {klokke} Error loading clock >>error_log.txt")
        driver.quit()
        sys.exit()
def add_tall(liste,maskin_nr,x,y,y1):
    try:
        excel,excel_ark,excel_ark1 = open_excel()
    except:
        klokke,dato = dato_tid()
        os.system(f"cd {error_log_path} && echo {dato} {klokke} Error failed to load excel >>error_log.txt")
        driver.quit()
        sys.exit()
    try:    
        count=0
        count = excel_ark1.cell_value(y1,x)
        if count == "":
            count=0
        int(count)
        count+=1
        excel_ark.cell(y1+1,x+1,count)
        if liste[0] == "Idle" or liste[0] == "Closing":
            verdi=0
            verdi = excel_ark1.cell_value(y,x)
            if verdi == "":
                verdi=0
            int(verdi)
            verdi+=1
            excel_ark.cell(y+1,x+1,verdi)
        excel.save(excel_fil)
    except:
        klokke,dato = dato_tid()
        os.system(f"cd {error_log_path} && echo {dato} {klokke} Error writing to excel >>error_log.txt")
        driver.quit()
        sys.exit()
def open_webpage_and_run():
    try:
        maskin_liste = []
        link = "https://mielelogic.com/#/laundrystate"
        alle_maskiner =[
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[1]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[2]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[3]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[4]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[5]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[6]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[7]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[8]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[9]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[10]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[11]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[12]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[13]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[14]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[15]/div/div[5]',
            '//*[@id="content"]/div/div/div/div/div/div[1]/div/div[16]/div/div[5]'  
            ]
        
        driver.get(link)
        sleep(15)
        click_norge = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[1]/div/div/button')
        click_norge.click()
        click_norge = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[1]/div/div/ul/li[2]/a/span[1]')
        click_norge.click()
        brukernavn = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[2]/div/div/input')
        brukernavn.click()
        brukernavn.send_keys("My user name")
        passord = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[3]/div[1]/div[1]/input')
        passord.click()
        passord.send_keys("My password")
        login = driver.find_element_by_xpath('//*[@id="div1"]/form/div/div[2]/div[2]/div[4]/button')
        login.click()
        sleep(30)          
        vaskeri_status = driver.find_element_by_xpath('//*[@id="content"]/div/div/div/div/div[3]/div/div[1]/a/h4')
        vaskeri_status.click()
        klokke,dag = klokke_dag()
        sleep(30)         
        count=0
        driver.find_element_by_xpath('//*[@id="content"]/div/div/div/div/div/div[1]/div/div[1]/div/div[3]/span/i[2]')
    except:
        klokke,dato = dato_tid()
        os.system(f"cd {error_log_path} && echo {dato} {klokke} Error loading website pre loop >>error_log.txt")
        driver.quit()
        sys.exit()
    
    for e in alle_maskiner:    
        try:    
            for i in driver.find_elements_by_xpath(e):
                maskin_liste.append(i.text)
        except:
            klokke,dato = dato_tid()
            os.system(f"cd {error_log_path} && echo {dato} {klokke} Error looping through washingmachines >>error_log.txt")
            driver.quit()
            sys.exit()
        else:    
            maskin_1 = maskin_liste[count].split(" ")
            print(maskin_1[0])
            print("1")
            x,y,y1 = klokke_maskin(count,klokke,dag)
            add_tall(maskin_1,count,x,y,y1)
            count+=1


def error_count():
    excel,excel_ark,excel_ark1 = open_excel()
    count=0
    count = excel_ark1.cell_value(2,26)
    if count == "":
        count=0
    count+=1
    excel_ark.cell(3,27,count)
    excel.save(excel_fil)

    
    
def main():
    open_webpage_and_run()
    driver.quit()
    error_count()
    sys.exit()
main()